"""Multi-sheet, formatted Excel export (Phase 19)."""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules import queries

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NUMERIC_HEADERS = {
    "quantity", "qty", "total qty", "unit_price", "extended_price",
    "rfqs", "items", "vessels", "buyers", "rfq count", "lines", "count",
    "total_line_items",
}

_DOCUMENT_COLS = [
    "file_name", "document_format", "document_type", "rfq_no", "quotation_no",
    "buyer_name", "supplier_name", "vessel_name", "imo_number", "port_name",
    "department", "subject", "currency", "issued_date", "due_date",
    "deliver_by_date", "vessel_eta", "vessel_etd", "payment_terms",
    "total_line_items", "duplicate_flag", "extraction_status", "confidence_score",
    "created_at",
]
_LINE_ITEM_COLS = [
    "file_name", "rfq_no", "quotation_no", "buyer_name", "vessel_name",
    "imo_number", "port_name", "issued_date", "item_no", "section_name",
    "manufacturer", "model", "part_type", "part_number", "buyer_part_number",
    "item_code", "description", "normalized_item_name", "category", "uom",
    "quantity", "unit_price", "extended_price", "currency", "remarks",
]


def _subset(df, cols):
    return df[[c for c in cols if c in df.columns]] if not df.empty else df


def _format_sheet(ws, df):
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    numeric_idx = {
        i for i, c in enumerate(df.columns, start=1)
        if str(c).strip().lower() in _NUMERIC_HEADERS
    }
    for col_idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        sample = df[col].head(200).astype(str).map(len).max() if not df.empty else 0
        width = min(max(int(sample or 0), len(str(col))) + 2, 50)
        ws.column_dimensions[letter].width = max(width, 10)
        if col_idx in numeric_idx:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "#,##0.##"


def build_workbook() -> bytes:
    """Build the analytics workbook and return its bytes."""
    sheets = {
        "Documents": _subset(queries.load_documents(), _DOCUMENT_COLS),
        "Line_Items": _subset(queries.load_line_items(), _LINE_ITEM_COLS),
        "Vessel_Summary": queries.vessel_analysis(),
        "Item_Summary": queries.item_analysis(),
        "Buyer_Summary": queries.buyer_analysis(),
        "Port_Summary": queries.port_analysis(),
        "Review_Errors": _subset(
            queries.load_review_errors(),
            ["file_name", "rfq_no", "issue_type", "field_name", "extracted_value",
             "suggested_value", "confidence", "review_status", "created_at"],
        ),
    }
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None or df.empty:
                df = pd.DataFrame({"(no data)": []})
            df.to_excel(writer, sheet_name=name, index=False)
            _format_sheet(writer.sheets[name], df)
    buf.seek(0)
    return buf.getvalue()


def default_filename() -> str:
    return f"CargoPulse_Analytics_{datetime.now():%Y%m%d}.xlsx"


def write_to_file(path) -> str:
    with open(path, "wb") as f:
        f.write(build_workbook())
    return str(path)
